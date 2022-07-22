from django.shortcuts import render, redirect, HttpResponse
from app01 import models
from django.views.decorators.csrf import csrf_exempt
from django.core.validators import RegexValidator
from django.core.exceptions import ValidationError
from django import forms
from django.utils.safestring import mark_safe
from app01.utils.pagination import Pagination
from openpyxl import load_workbook


# Create your views here.
# 部门列表

def depart_list(request):
    """ 部门列表 """
    # info = request.session.get("info")
    # if not info:
    #     return redirect('/login/')
    # 去数据库中获取所有的部门信息，models.Department.objects.all()
    # [对象，对象，对象]
    # 获取所有用户列表[obj,obj]
    queryset = models.Department.objects.all()
    # 加分页
    page_object = Pagination(request, queryset, page_size=2)
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.html(),
    }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    """ 添加部门 """
    if request.method == "GET":
        return render(request, 'depart_add.html')

    # 获取用户POST提交过来的数据 (title输入为空怎么办？）
    title = request.POST.get("title")
    # 保存到数据库
    models.Department.objects.create(title=title)
    # 重定向到部门列表页面
    return redirect('/depart/list/')


def depart_delete(request):
    """ 删除 """
    # https://127.0.0.1:8000/depart/delete/?nid=1
    # 获取ID
    nid = request.GET.get("nid")

    # 删除
    models.Department.objects.filter(id=nid).delete()
    # 跳转回部门列表
    return redirect('/depart/list')


# http://127.0.0.1:8000/depart/10/edit/
# http://127.0.0.1:8000/depart/2/edit/
# http://127.0.0.1:8000/depart/1/edit/
def depart_edit(request, nid):
    """ 修改部门 """
    if request.method == "GET":
        # 根据nid，获取他的数据[obj,]
        row_object = models.Department.objects.filter(id=nid).first()

        return render(request, 'depart_edit.html', {"row_object": row_object})
    # 获取用户提交的数据
    title = request.POST.get("title")
    # 根据ID找到数据库中的数据进行更新
    models.Department.objects.filter(id=nid).update(title=title)
    # 重定向回部门列表
    return redirect("/depart/list/")


@csrf_exempt
def depart_multi(request):
    """ 批量上传，基于excel文件 """
    # 1.获取用户上传的 文件对象
    file_object = request.FILES.get("exc")
    # <class 'django.core.files.uploadedfile.InMemoryUploadedFile'>
    # print(type(file_object))

    # 2.对象传递给openpyxl，有openpyxl读取文件的内容

    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # # <class 'django.core.files.uploadedfile.InMemoryUploadedFile'>
    # 12
    #
    # cell = sheet.cell(1, 1)
    # print(cell.value)
    # 3. 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    # 直接打开excel并读取内容
    return redirect('/depart/list/')
